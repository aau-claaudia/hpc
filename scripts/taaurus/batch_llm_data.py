#!/usr/bin/env python3
"""
Prepare and split JSONL input for TAAURUS batch LLM jobs.

Each input line must be a JSON object with:
  - "id"   : unique identifier (string or number)
  - "text" : document or prompt text to send to the model

No third-party dependencies (stdlib only).
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Any, Iterable, Iterator


REQUIRED_FIELDS = ("id", "text")


def _normalize_record(raw: dict[str, Any], line_no: int, source: str) -> dict[str, Any]:
    if "text" not in raw:
        raise ValueError(f"{source} line {line_no}: missing required field 'text'")
    text = raw["text"]
    if not isinstance(text, str):
        raise ValueError(f"{source} line {line_no}: 'text' must be a string")
    text = text.strip()
    if not text:
        raise ValueError(f"{source} line {line_no}: 'text' must not be empty")

    if "id" in raw:
        record_id = raw["id"]
    else:
        record_id = line_no

    return {"id": record_id, "text": text}


def _write_jsonl(records: Iterable[dict[str, Any]], path: Path) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    with path.open("w", encoding="utf-8") as fh:
        for record in records:
            fh.write(json.dumps(record, ensure_ascii=False) + "\n")
            count += 1
    return count


def _read_jsonl(path: Path) -> Iterator[dict[str, Any]]:
    with path.open(encoding="utf-8") as fh:
        for line_no, line in enumerate(fh, start=1):
            line = line.strip()
            if not line:
                continue
            try:
                raw = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(f"{path} line {line_no}: invalid JSON ({exc})") from exc
            if not isinstance(raw, dict):
                raise ValueError(f"{path} line {line_no}: each line must be a JSON object")
            yield _normalize_record(raw, line_no, str(path))


def _records_from_csv(
    path: Path,
    id_column: str,
    text_column: str,
    *,
    has_header: bool,
) -> Iterator[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8") as fh:
        if has_header:
            reader = csv.DictReader(fh)
            if not reader.fieldnames:
                raise ValueError(f"{path}: CSV has no header row")
            for name in (id_column, text_column):
                if name not in reader.fieldnames:
                    raise ValueError(
                        f"{path}: column '{name}' not found. "
                        f"Available: {', '.join(reader.fieldnames)}"
                    )
            for line_no, row in enumerate(reader, start=2):
                yield _normalize_record(
                    {"id": row[id_column], "text": row[text_column]},
                    line_no,
                    str(path),
                )
        else:
            reader = csv.reader(fh)
            col_index = {id_column: 0, text_column: 1}
            if id_column not in col_index or text_column not in col_index:
                try:
                    id_idx = int(id_column)
                    text_idx = int(text_column)
                except ValueError as exc:
                    raise ValueError(
                        "Without --csv-header, use --id-column and --text-column "
                        "as 0-based column indices (e.g. 0 and 1)"
                    ) from exc
                col_index = {id_column: id_idx, text_column: text_idx}
            for line_no, row in enumerate(reader, start=1):
                if len(row) <= max(col_index.values()):
                    raise ValueError(f"{path} line {line_no}: not enough columns")
                yield _normalize_record(
                    {
                        "id": row[col_index[id_column]],
                        "text": row[col_index[text_column]],
                    },
                    line_no,
                    str(path),
                )


def _records_from_txt_dir(directory: Path) -> Iterator[dict[str, Any]]:
    files = sorted(p for p in directory.iterdir() if p.is_file() and p.suffix.lower() == ".txt")
    if not files:
        raise ValueError(f"{directory}: no .txt files found")
    for path in files:
        text = path.read_text(encoding="utf-8").strip()
        if not text:
            raise ValueError(f"{path}: file is empty")
        yield {"id": path.stem, "text": text}


def _records_from_xlsx(path: Path, id_column: str, text_column: str, sheet: str | int) -> Iterator[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "Excel support requires openpyxl. Install with: python -m pip install --user openpyxl\n"
            "Or export the sheet to CSV in Excel/LibreOffice and use --source csv."
        ) from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    if isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    else:
        if sheet not in wb.sheetnames:
            raise ValueError(f"{path}: sheet '{sheet}' not found. Available: {', '.join(wb.sheetnames)}")
        ws = wb[sheet]

    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    except StopIteration:
        raise ValueError(f"{path}: worksheet is empty") from None

    for name in (id_column, text_column):
        if name not in header:
            raise ValueError(f"{path}: column '{name}' not found. Available: {', '.join(header)}")

    id_idx = header.index(id_column)
    text_idx = header.index(text_column)
    for line_no, row in enumerate(rows, start=2):
        if row is None or all(cell is None for cell in row):
            continue
        values = list(row)
        if max(id_idx, text_idx) >= len(values):
            continue
        yield _normalize_record(
            {"id": values[id_idx], "text": values[text_idx]},
            line_no,
            str(path),
        )


def _extract_pdf_text(path: Path) -> str:
    _require_command("pdftotext")
    result = subprocess.run(
        ["pdftotext", "-enc", "UTF-8", str(path), "-"],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        check=False,
    )
    if result.returncode != 0:
        raise ValueError(f"{path}: pdftotext failed ({result.stderr.strip()})")
    return result.stdout.strip()


def _extract_docx_text(path: Path) -> str:
    try:
        from docx import Document
    except ImportError as exc:
        raise SystemExit(
            "Word (.docx) support requires python-docx. Install with: "
            "python -m pip install --user python-docx\n"
            "Or convert to .txt with LibreOffice (see the batch LLM guide)."
        ) from exc

    doc = Document(path)
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    if not paragraphs:
        raise ValueError(f"{path}: no text found in document")
    return "\n\n".join(paragraphs)


def _require_command(name: str) -> None:
    if shutil.which(name) is None:
        raise SystemExit(
            f"Required command '{name}' was not found in PATH.\n"
            "Install it on the login node, or convert files manually (see the batch LLM guide)."
        )


def _records_from_pdf_dir(directory: Path) -> Iterator[dict[str, Any]]:
    files = sorted(p for p in directory.iterdir() if p.is_file() and p.suffix.lower() == ".pdf")
    if not files:
        raise ValueError(f"{directory}: no .pdf files found")
    for path in files:
        text = _extract_pdf_text(path)
        if not text:
            raise ValueError(f"{path}: extracted text is empty")
        yield {"id": path.stem, "text": text}


def _records_from_docx_dir(directory: Path) -> Iterator[dict[str, Any]]:
    files = sorted(p for p in directory.iterdir() if p.is_file() and p.suffix.lower() == ".docx")
    if not files:
        raise ValueError(f"{directory}: no .docx files found")
    for path in files:
        yield {"id": path.stem, "text": _extract_docx_text(path)}


def _records_from_txt_lines(path: Path) -> Iterator[dict[str, Any]]:
    with path.open(encoding="utf-8") as fh:
        for line_no, line in enumerate(fh, start=1):
            text = line.strip()
            if not text:
                continue
            yield {"id": line_no, "text": text}


def _parse_sheet(value: str) -> str | int:
    try:
        return int(value)
    except ValueError:
        return value


def cmd_prepare(args: argparse.Namespace) -> int:
    source = args.source
    sheet = _parse_sheet(str(args.sheet))
    if source == "jsonl":
        records = list(_read_jsonl(args.input))
    elif source == "csv":
        records = list(
            _records_from_csv(
                args.input,
                args.id_column,
                args.text_column,
                has_header=args.csv_header,
            )
        )
    elif source == "txt-lines":
        records = list(_records_from_txt_lines(args.input))
    elif source == "txt-dir":
        records = list(_records_from_txt_dir(args.input))
    elif source == "xlsx":
        records = list(
            _records_from_xlsx(
                args.input,
                args.id_column,
                args.text_column,
                sheet,
            )
        )
    elif source == "pdf-dir":
        records = list(_records_from_pdf_dir(args.input))
    elif source == "docx-dir":
        records = list(_records_from_docx_dir(args.input))
    else:
        raise ValueError(f"unknown source: {source}")

    if not records:
        print("No records to write.", file=sys.stderr)
        return 1

    count = _write_jsonl(records, args.output)
    print(f"Wrote {count} records to {args.output}")
    return 0


def cmd_validate(args: argparse.Namespace) -> int:
    errors = 0
    count = 0
    seen_ids: set[Any] = set()
    for record in _read_jsonl(args.input):
        count += 1
        record_id = record["id"]
        if record_id in seen_ids:
            print(f"duplicate id: {record_id!r}", file=sys.stderr)
            errors += 1
        seen_ids.add(record_id)
    if count == 0:
        print(f"{args.input}: no records found", file=sys.stderr)
        return 1
    if errors:
        print(f"Validation failed: {errors} issue(s) in {count} records", file=sys.stderr)
        return 1
    print(f"OK: {count} records, required fields present, ids unique")
    return 0


def cmd_split(args: argparse.Namespace) -> int:
    records = list(_read_jsonl(args.input))
    if not records:
        print(f"{args.input}: no records to split", file=sys.stderr)
        return 1

    num_chunks = args.num_chunks
    if num_chunks < 1:
        print("--num-chunks must be at least 1", file=sys.stderr)
        return 1
    if num_chunks > len(records):
        print(
            f"Warning: --num-chunks ({num_chunks}) > record count ({len(records)}); "
            f"using {len(records)} chunks",
            file=sys.stderr,
        )
        num_chunks = len(records)

    chunk_size = math.ceil(len(records) / num_chunks)
    pad = max(3, len(str(num_chunks - 1)))
    out_dir = args.output
    out_dir.mkdir(parents=True, exist_ok=True)

    written = 0
    for index in range(num_chunks):
        start = index * chunk_size
        end = min(start + chunk_size, len(records))
        if start >= len(records):
            break
        chunk_path = out_dir / f"chunk_{index:0{pad}d}.jsonl"
        n = _write_jsonl(records[start:end], chunk_path)
        written += n
        print(f"  {chunk_path.name}: {n} records")

    print(f"Split {len(records)} records into {index + 1} chunks under {out_dir}")
    if written != len(records):
        print("Internal error: record count mismatch", file=sys.stderr)
        return 1
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Prepare JSONL input and chunk files for TAAURUS batch LLM jobs.",
    )
    sub = parser.add_subparsers(dest="command", required=True)

    prepare = sub.add_parser(
        "prepare",
        help="Convert CSV, text files, or JSONL into standard input.jsonl",
    )
    prepare.add_argument(
        "--source",
        choices=("jsonl", "csv", "xlsx", "txt-lines", "txt-dir", "pdf-dir", "docx-dir"),
        required=True,
        help=(
            "Input format: jsonl, csv, xlsx (needs openpyxl), txt-lines, txt-dir, "
            "pdf-dir (needs pdftotext), docx-dir (needs python-docx)"
        ),
    )
    prepare.add_argument("--input", type=Path, required=True, help="Input file or directory")
    prepare.add_argument(
        "--output",
        type=Path,
        required=True,
        help="Output JSONL path (e.g. /media/projectA/data/input.jsonl)",
    )
    prepare.add_argument(
        "--id-column",
        default="id",
        help="CSV: column name (with --csv-header) or 0-based index (without header). Default: id",
    )
    prepare.add_argument(
        "--text-column",
        default="text",
        help="CSV: column name or 0-based index. Default: text",
    )
    prepare.add_argument(
        "--csv-header",
        action="store_true",
        help="CSV: first row contains column names",
    )
    prepare.add_argument(
        "--sheet",
        default=0,
        help="Excel: worksheet name or 0-based index (default: 0)",
    )
    prepare.set_defaults(func=cmd_prepare)

    validate = sub.add_parser("validate", help="Check that a JSONL file is ready for batch jobs")
    validate.add_argument("--input", type=Path, required=True)
    validate.set_defaults(func=cmd_validate)

    split = sub.add_parser("split", help="Split input.jsonl into chunk files for Slurm array jobs")
    split.add_argument("--input", type=Path, required=True)
    split.add_argument("--output", type=Path, required=True, help="Output directory for chunks")
    split.add_argument(
        "--num-chunks",
        type=int,
        required=True,
        help="Number of chunk files (should match Slurm array size)",
    )
    split.set_defaults(func=cmd_split)

    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
